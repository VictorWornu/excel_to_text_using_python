from openpyxl import load_workbook
from datetime import date, datetime

workbook = load_workbook(filename = 'main.xlsx')
worksheet = workbook.active

file = open("main.txt", "a")

todays_date = date.today()
current_time = datetime.now()
current_time_upd = current_time.strftime("%H:%M:%S")
time_updated = (f"Updated {current_time_upd} , {todays_date}")
time_updated_upd = time_updated + "\n"

file.write(time_updated_upd)

last_row = worksheet.max_row


for row in range(2,last_row+1):
	id_location = ("A" + str(row))
	firstname_location = ("B" + str(row))
	surname_location = ("C" + str(row))
	year_location = ("D" + str(row))
	course_location = ("E" + str(row))
	id_no = worksheet[id_location].value
	firstname = worksheet[firstname_location].value
	surname	= worksheet[surname_location].value
	year = worksheet[year_location]. value
	course = worksheet[course_location].value
	data_extracted = (f"{id_no}. {firstname} {surname} is in year {year} in the {course} department.")
	data_extracted_upd = data_extracted + "\n"
	file.write(data_extracted_upd)
new_line = '\n'
file.write(new_line)
file.close()